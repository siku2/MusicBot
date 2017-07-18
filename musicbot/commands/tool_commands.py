import re

from discord import Embed
from openpyxl import Workbook

from ..bookmarks import bookmark
from ..logger import OnlineLogger
from ..random_sets import RandomSets
from ..reminder import Action, Calendar
from ..settings import Settings
from ..utils import (Response, block_user, clean_songname, command_info,
                     owner_only, parse_timestamp, to_timestamp)


class ToolCommands:
    @block_user
    async def cmd_reminder(self, channel, author, player, server, leftover_args):
        """
        Usage:
            ***REMOVED***command_prefix***REMOVED***reminder create
            ***REMOVED***command_prefix***REMOVED***reminder list

        Create a reminder!
        """

        if len(leftover_args) < 1:
            return Response("Please git gud!")

        command = leftover_args[0].lower().strip()

        if (command == "create"):
            import parsedatetime
            cal = parsedatetime.Calendar()

            reminder_name = None
            reminder_due = None
            reminder_repeat = None
            reminder_end = None
            reminder_action = None

            # find out the name
            def check(m):
                return len(m.content) > 3

            msg = await self.safe_send_message(
                channel, "How do you want to call your reminder?")
            response = await self.wait_for_message(
                author=author, channel=channel, check=check)
            reminder_name = response.content
            await self.safe_delete_message(msg)
            await self.safe_delete_message(response)

            # find out the due date
            while True:
                msg = await self.safe_send_message(channel, "When is it due?")
                response = await self.wait_for_message(
                    author=author, channel=channel)

                reminder_due = datetime(
                    *cal.parse(response.content.strip().lower())[0][:6])
                await self.safe_delete_message(msg)
                if reminder_due is not None:
                    await self.safe_delete_message(response)
                    break

                await self.safe_delete_message(response)

            # repeated reminder
            while True:
                msg = await self.safe_send_message(
                    channel,
                    "When should this reminder be repeated? (\"never\" if not at all)"
                )
                response = await self.wait_for_message(
                    author=author, channel=channel)
                await self.safe_delete_message(msg)
                if (response.content.lower().strip() in ("n", "no", "nope",
                                                         "never")):
                    await self.safe_delete_message(response)
                    reminder_repeat = None
                    break

                reminder_repeat = datetime(*cal.parse(
                    response.content.strip().lower())[0][:6]) - datetime.now()
                if reminder_repeat is not None:
                    await self.safe_delete_message(response)
                    break

                await self.safe_delete_message(response)

            # reminder end
            if reminder_repeat is not None:
                while True:
                    msg = await self.safe_send_message(
                        channel,
                        "When should this reminder stop being repeated? (\"never\" if not at all)"
                    )
                    response = await self.wait_for_message(
                        author=author, channel=channel)
                    await self.safe_delete_message(msg)
                    if (response.content.lower().strip() in ("n", "no", "nope",
                                                             "never")):
                        await self.safe_delete_message(response)
                        reminder_end = None
                        break

                    reminder_end = datetime(
                        *cal.parse(response.content.strip().lower())[0][:6])
                    if reminder_end is not None:
                        await self.safe_delete_message(response)
                        break

                    await self.safe_delete_message(response)

            # action
            def check(m):
                try:
                    if 4 > int(m.content) > 0:
                        return True
                    else:
                        return False
                except:
                    return False

            selected_action = 0

            while True:
                msg = await self.safe_send_message(
                    channel,
                    "**Select one:**\n```\n1: Send a message\n2: Play a video\n3: Play an alarm sound```"
                )
                response = await self.wait_for_message(
                    author=author, channel=channel)
                await self.safe_delete_message(msg)
                selected_action = int(response.content)

                if selected_action is not None:
                    await self.safe_delete_message(response)
                    break

                await self.safe_delete_message(response)

            # action 1 (message)
            if selected_action == 1:
                action_message = "Your reminder *****REMOVED***reminder.name***REMOVED***** is due"
                action_channel = None
                action_delete_after = 0
                action_delete_previous = False

                # find message
                msg = await self.safe_send_message(
                    channel, "What should the message say?")
                response = await self.wait_for_message(
                    author=author, channel=channel)
                action_message = response.content
                await self.safe_delete_message(msg)
                await self.safe_delete_message(response)

                # find channel
                while action_channel is None:
                    msg = await self.safe_send_message(
                        channel,
                        "To which channel should the message be sent?\nPossible inputs:\n\n:white_small_square: Channel id or channel name\n:white_small_square: \"me\" for a private message\n:white_small_square: \"this\" to select the current channel\n:white_small_square: You can also @mention people or #mention a channel"
                    )
                    response = await self.wait_for_message(
                        author=author, channel=channel)

                    if len(response.channel_mentions) > 0:
                        action_channel = response.channel_mentions[0]
                    elif len(response.mentions) > 0:
                        action_channel = response.mentions[0]
                    elif response.content.lower().strip() == "me":
                        action_channel = author
                    elif response.content.lower().strip() == "this":
                        action_channel = channel
                    else:
                        return Response("not yet implemented :P")

                    await self.safe_delete_message(msg)
                    await self.safe_delete_message(response)

                # find delete after time
                def check(m):
                    try:
                        if m.content.lower().strip() in [
                                "never", "no"
                        ] or int(m.content.strip()) >= 0:
                            return True
                        else:
                            return False
                    except:
                        return False

                msg = await self.safe_send_message(
                    channel,
                    "After how many seconds should the message be deleted? (\"never\" for not at all)"
                )
                response = await self.wait_for_message(
                    author=author, channel=channel, check=check)
                if response.content.lower().strip() in ["never", "no"]:
                    action_delete_after = 0
                else:
                    action_delete_after = int(response.content.strip())

                await self.safe_delete_message(msg)
                await self.safe_delete_message(response)

                # find if delete old message
                if reminder_repeat is not None:
                    msg = await self.safe_send_message(
                        channel,
                        "Before sending a new message, should the old one be deleted?"
                    )
                    response = await self.wait_for_message(
                        author=author, channel=channel)
                    if response.content.lower().strip() in ["y", "yes"]:
                        action_delete_previous = True

                    await self.safe_delete_message(msg)
                    await self.safe_delete_message(response)

                reminder_action = Action(
                    channel=action_channel,
                    msg_content=action_message,
                    delete_msg_after=action_delete_after,
                    delete_old_message=action_delete_previous)

            # action 2 (play url)
            elif selected_action == 2:
                action_source_url = ""
                action_voice_channel = None

                # find video url
                msg = await self.safe_send_message(
                    channel, "What's the url of the video you want to play?")
                response = await self.wait_for_message(
                    author=author, channel=channel)
                action_source_url = response.content
                await self.safe_delete_message(msg)
                await self.safe_delete_message(response)

                # find playback channel
                msg = await self.safe_send_message(
                    channel,
                    "To which channel should the video be played?\nPossible inputs:\n\n:white_small_square: Channel id or channel name\n:white_small_square: \"this\" to select your current channel"
                )
                response = await self.wait_for_message(
                    author=author, channel=channel)

                if response.content.lower().strip() == "this":
                    return Response("not yet implemented :P")
                else:
                    return Response("not yet implemented :P")

            # action 3 (play predefined)
            elif selected_action == 3:
                pass

            # finalizing
            self.calendar.create_reminder(
                reminder_name,
                reminder_due,
                reminder_action,
                repeat_every=reminder_repeat,
                repeat_end=reminder_end)
            return Response(
                "Created a reminder called *****REMOVED******REMOVED*****\ndue: ***REMOVED******REMOVED***\nrepeat: ***REMOVED******REMOVED***\nrepeat end: ***REMOVED******REMOVED***\naction: ***REMOVED******REMOVED***".
                format(reminder_name, reminder_due, reminder_repeat,
                       reminder_end, reminder_action))

        elif (command == "list"):
            if len(self.calendar.reminders) < 1:
                return Response("There are no reminders")

            text = ""
            for reminder in self.calendar.reminders:
                text += "****REMOVED***.name***REMOVED****\n".format(reminder)

            return Response(text)

    @command_info("2.0.3", 1485516420, ***REMOVED***
        "3.7.5": (1481827320, "The command finally works like it should"),
        "3.9.9": (1499977057, "moving Giesela too")
    ***REMOVED***)
    async def cmd_moveus(self, channel, server, author, message, leftover_args):
        """
        ///|Usage
        `***REMOVED***command_prefix***REMOVED***moveus <channel name>`
        ///|Explanation
        Move everyone in your current channel to another one!
        """

        if len(leftover_args) < 1:
            return Response("You need to provide a target channel")

        search_channel = " ".join(leftover_args)
        if search_channel.lower().strip() == "home":
            search_channel = "Giesela's reign"

        if author.voice.voice_channel is None:
            return Response(
                "You're incredibly incompetent to do such a thing!")

        author_channel = author.voice.voice_channel
        voice_members = author_channel.voice_members
        move_myself = False
        if server.me in voice_members:
            voice_members.remove(server.me)
            move_myself = True

        target_channel = self.get_channel(search_channel)
        if target_channel is None:
            for chnl in server.channels:
                if chnl.name == search_channel and chnl.type == ChannelType.voice:
                    target_channel = chnl
                    break

        if target_channel is None:
            return Response(
                "Can't resolve the target channel!")

        s = 0
        for voice_member in voice_members:
            await self.move_member(voice_member, target_channel)
            s += 1

        if move_myself:
            print("moving myself")
            await self.move_voice_client(target_channel)

    async def cmd_summon(self, channel, author, voice_channel):
        """
        Usage:
            ***REMOVED***command_prefix***REMOVED***summon

        Call the bot to the summoner's voice channel.
        """

        if not author.voice_channel:
            raise exceptions.CommandError('You are not in a voice channel!')

        voice_client = self.the_voice_clients.get(channel.server.id, None)
        if voice_client and voice_client.channel.server == author.voice_channel.server:
            await self.move_voice_client(author.voice_channel)
            return

        # move to _verify_vc_perms?
        chperms = author.voice_channel.permissions_for(
            author.voice_channel.server.me)

        if not chperms.connect:
            print("Cannot join channel \"%s\", no permission." %
                  author.voice_channel.name)
            return Response(
                "```Cannot join channel \"%s\", no permission.```" %
                author.voice_channel.name,
                delete_after=25)

        elif not chperms.speak:
            print("Will not join channel \"%s\", no permission to speak." %
                  author.voice_channel.name)
            return Response(
                "```Will not join channel \"%s\", no permission to speak.```" %
                author.voice_channel.name,
                delete_after=25)

        player = await self.get_player(author.voice_channel, create=True)

        if player.is_stopped:
            player.play()

        if self.config.auto_playlist:
            await self.on_player_finished_playing(player)

    @command_info("1.0.0", 1477180800, ***REMOVED***
        "2.0.2": (1481827560, "Can now use @mentions to \"goto\" a user")
    ***REMOVED***)
    async def cmd_goto(self, server, channel, user_mentions, author, leftover_args):
        """
        Usage:
            ***REMOVED***command_prefix***REMOVED***goto <id | name | @mention>

        Call the bot to a channel.
        """

        channelID = " ".join(leftover_args)
        if channelID.lower() == "home":
            await self.goto_home(server)
            return Response("yep")

        targetChannel = self.get_channel(channelID)
        if targetChannel is None:
            for chnl in server.channels:
                if chnl.name == channelID and chnl.type == ChannelType.voice:
                    targetChannel = chnl
                    break
            else:
                if user_mentions:
                    for ch in server.channels:
                        for user in ch.voice_members:
                            if user in user_mentions:
                                targetChannel = ch
                    if targetChannel is None:
                        return Response(
                            "Cannot find *****REMOVED******REMOVED***** in any voice channel".format(
                                ", ".join([x.mention for x in user_mentions])))
                else:
                    print("Cannot find channel \"%s\"" % channelID)
                    return Response(
                        "```Cannot find channel \"%s\"```" % channelID)

        voice_client = await self.get_voice_client(targetChannel)
        print("Will join channel \"%s\"" % targetChannel.name)
        await self.move_voice_client(targetChannel)

        # move to _verify_vc_perms?
        chperms = targetChannel.permissions_for(targetChannel.server.me)

        if not chperms.connect:
            print("Cannot join channel \"%s\", no permission." %
                  targetChannel.name)
            return Response(
                "```Cannot join channel \"%s\", no permission.```" %
                targetChannel.name)

        elif not chperms.speak:
            print("Will not join channel \"%s\", no permission to speak." %
                  targetChannel.name)
            return Response(
                "```Will not join channel \"%s\", no permission to speak.```" %
                targetChannel.name)

        player = await self.get_player(targetChannel, create=True)

        if player.is_stopped:
            player.play()

        if self.config.auto_playlist:
            await self.on_player_finished_playing(player)

        return Response("Joined the channel *****REMOVED******REMOVED*****".format(targetChannel.name))

    @owner_only
    async def cmd_countmsgs(self, server, author, channel_id, number):
        alphabet = list("abcdefghijklmnopqrstuvwxyz")

        def index_to_alphabet(ind):
            if ind < len(alphabet):
                return alphabet[ind].upper()

            remainder = ind % len(alphabet)
            return index_to_alphabet(ind -
                                     remainder) + alphabet[remainder].upper()

        msgs_by_member = ***REMOVED******REMOVED***
        msgs_by_date = OrderedDict()
        answers_by_date = OrderedDict()
        channel = server.get_channel(channel_id)
        last_msg = None
        last_answer = None
        spam = 0

        async for msg in self.logs_from(channel, limit=int(number)):
            increment = 1
            if last_msg is not None and msg.author.id == last_msg.author.id and abs(
                    (last_msg.timestamp - msg.timestamp).total_seconds()) < 10:
                spam += 1
                last_msg = msg
                increment = 0

            if last_answer is None or last_answer.author != msg.author:
                dt = answers_by_date.get(
                    "***REMOVED***0.day:0>2***REMOVED***/***REMOVED***0.month:0>2***REMOVED***/***REMOVED***0.year:0>4***REMOVED***".format(
                        msg.timestamp), ***REMOVED******REMOVED***)
                dt[msg.author.id] = dt.get(msg.author.id, 0) + increment
                answers_by_date["***REMOVED***0.day:0>2***REMOVED***/***REMOVED***0.month:0>2***REMOVED***/***REMOVED***0.year:0>4***REMOVED***".
                                format(msg.timestamp)] = dt
                last_answer = msg

            existing_msgs = msgs_by_member.get(msg.author.id, [0, 0])
            existing_msgs[0] += increment
            existing_msgs[1] += len(re.sub(r"\W", r"", msg.content))
            msgs_by_member[msg.author.id] = existing_msgs
            dt = msgs_by_date.get(
                "***REMOVED***0.day:0>2***REMOVED***/***REMOVED***0.month:0>2***REMOVED***/***REMOVED***0.year:0>4***REMOVED***".format(msg.timestamp),
                ***REMOVED******REMOVED***)
            dt[msg.author.id] = dt.get(msg.author.id, 0) + increment
            msgs_by_date["***REMOVED***0.day:0>2***REMOVED***/***REMOVED***0.month:0>2***REMOVED***/***REMOVED***0.year:0>4***REMOVED***".format(
                msg.timestamp)] = dt
            last_msg = msg

        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"
        ws2 = wb.create_sheet("Answers")
        ws["A2"] = "TOTAL"
        sorted_user_index = ***REMOVED******REMOVED***
        i = 1
        for member in sorted(msgs_by_member):
            data = msgs_by_member[member]
            ws["***REMOVED******REMOVED******REMOVED******REMOVED***".format("A", i)] = server.get_member(
                member
            ).name if server.get_member(member) is not None else "Unknown"
            ws["***REMOVED******REMOVED******REMOVED******REMOVED***".format("B", i)] = data[0]
            ws["***REMOVED******REMOVED******REMOVED******REMOVED***".format("C", i)] = data[1]
            sorted_user_index[member] = index_to_alphabet(i)
            i += 1

        i += 1
        for date in reversed(msgs_by_date.keys()):
            ws["A" + str(i)] = date
            for mem in msgs_by_date[date]:
                ws["***REMOVED******REMOVED******REMOVED******REMOVED***".format(sorted_user_index.get(mem),
                                 i)] = msgs_by_date[date][mem]
            i += 1

        i = 1
        for date in reversed(answers_by_date.keys()):
            ws2["A" + str(i)] = date
            for mem in answers_by_date[date]:
                ws2["***REMOVED******REMOVED******REMOVED******REMOVED***".format(sorted_user_index.get(mem),
                                  i)] = answers_by_date[date][mem]
            i += 1

        wb.save("cache/last_data.xlsx")

        await self.send_file(
            author,
            open("cache/last_data.xlsx", "rb"),
            filename='%s-msgs.xlsx' % (server.name.replace(' ', '_')))

    async def cmd_archivechat(self, server, author, message, placeholder=None, number=1000000):
        if message.channel_mentions is None or len(
                message.channel_mentions) < 1:
            return Response("Stupid duck")

        channel = message.channel_mentions[0]
        msgs = []
        async for msg in self.logs_from(channel, limit=int(number)):
            msg_data = ***REMOVED***
                "name": msg.author.name,
                "timestamp": str(round(msg.timestamp.timestamp())),
                "content": msg.content,
                "attachments": msg.attachments
            ***REMOVED***
            msgs.append(msg_data)

        json.dump(msgs[::-1], open("cache/last_message_archive.json", "w+"))
        await self.send_file(
            author,
            open("cache/last_message_archive.json", "rb"),
            filename='%s-msg-archive.json' % (server.name.replace(' ', '_')))

    @owner_only
    async def cmd_surveyserver(self, server):
        if self.online_loggers.get(server.id, None) is not None:
            return Response("I'm already looking at this server")
        else:
            online_logger = OnlineLogger(self)
            self.online_loggers[server.id] = online_logger
            Settings["online_loggers"] = list(self.online_loggers.keys())
            return Response("okay, okay!")

    def load_online_loggers(self):
        for server_id in Settings.get_setting("online_loggers", default=[]):
            online_logger = OnlineLogger(self)
            self.online_loggers[server_id] = online_logger
            for listener in Settings.get_setting(
                    "online_logger_listeners_" + server_id, default=[]):
                online_logger.add_listener(listener)

    @owner_only
    async def cmd_evalsurvey(self, server, author):
        online_logger = self.online_loggers.get(server.id, None)
        if online_logger is None:
            return Response("I'm not even spying here")
        online_logger.create_output()
        await self.send_file(
            author,
            open("cache/last_survey_data.xlsx", "rb"),
            filename='%s-survey.xlsx' % (server.name.replace(' ', '_')))
        return Response("There you go, fam")

    @owner_only
    async def cmd_resetsurvey(self, server):
        online_logger = self.online_loggers.get(server.id, None)
        if online_logger is None:
            return Response("I'm not even spying here")
        online_logger.reset()
        return Response("Well then")

    async def cmd_notifyme(self, server, author):
        """
        Usage:
            ***REMOVED***command_prefix***REMOVED***notifyme

        Get notified when someone starts playing
        """
        online_logger = self.online_loggers.get(server.id, None)
        if online_logger is None:
            return Response("I'm not even spying here")
        if online_logger.add_listener(author.id):
            Settings["online_logger_listeners_" + server.id] = [
                *Settings.get_setting(
                    "online_logger_listeners_" + server.id, default=[]),
                author.id
            ]
            return Response("Got'cha!")
        else:
            try:
                Settings["online_logger_listeners_" + server.id] = [
                    x
                    for x in Settings.get_setting(
                        "online_logger_listeners_" + server.id, default=[])
                    if x != author.id
                ]
            except ValueError:
                pass

            return Response("Nevermore you shall be annoyed!")

    @command_info("2.2.1", 1493757540, ***REMOVED***
        "3.7.8": (1499019245, "Fixed quoting by content.")
    ***REMOVED***)
    async def cmd_quote(self, author, channel, message, leftover_args):
        """
        ///|Usage
        `***REMOVED***command_prefix***REMOVED***quote [#channel] <message id> [message id...]`
        `***REMOVED***command_prefix***REMOVED***quote [#channel] [@mention] \"<message content>\"`
        ///|Explanation
        Quote a message
        """

        quote_to_channel = channel
        target_author = None

        if message.channel_mentions:
            channel = message.channel_mentions[0]
            leftover_args = leftover_args[1:]

        if message.mentions:
            target_author = message.mentions[0]
            leftover_args = leftover_args[1:]

        if len(leftover_args) < 1:
            return Response("Please specify the message you want to quote")

        message_content = " ".join(leftover_args)
        if (message_content[0] == "\"" and message_content[-1] == "\"") or re.search(r"\D", message_content) is not None:
            message_content = message_content.replace("\"", "")
            async for msg in self.logs_from(channel, limit=3000):
                if msg.id != message.id and message_content.lower().strip() in msg.content.lower().strip():
                    if target_author is None or target_author.id == msg.author.id:
                        leftover_args = [msg.id, ]
                        break
            else:
                if target_author is not None:
                    return Response("Didn't find a message with that content from ***REMOVED******REMOVED***".format(target_author.mention))
                else:
                    return Response("Didn't find a message with that content")

        await self.safe_delete_message(message)
        for message_id in leftover_args:
            try:
                quote_message = await self.get_message(channel, message_id)
            except:
                return Response("Didn't find a message with the id `***REMOVED******REMOVED***`".
                                format(message_id))

            author_data = ***REMOVED***
                "name": quote_message.author.display_name,
                "icon_url": quote_message.author.avatar_url
            ***REMOVED***
            embed_data = ***REMOVED***
                "description": quote_message.content,
                "timestamp": quote_message.timestamp,
                "colour": quote_message.author.colour
            ***REMOVED***
            em = Embed(**embed_data)
            em.set_author(**author_data)
            await self.send_message(quote_to_channel, embed=em)
        return

    @command_info("3.2.5", 1496428380, ***REMOVED***
        "3.3.9": (1497521393, "Added edit sub-command"),
        "3.4.1": (1497550771, "Added the filter \"mine\" to the listing function"),
        "3.4.6": (1497617827, "when listing bookmarks, they musn't be \"inline\"."),
        "3.5.8": (1497827057, "Editing bookmarks now works as expected")
    ***REMOVED***)
    async def cmd_bookmark(self, author, player, leftover_args):
        """
        ///|Creation
        ***REMOVED***command_prefix***REMOVED***bookmark [name] [timestamp]
        ///|Explanation
        Create a new bookmark for the current entry. If no name is provided the entry's title will be used and if there's no timestamp provided the current timestamp will be used.
        ///|Using
        ***REMOVED***command_prefix***REMOVED***bookmark <id | name>
        ///|Editing
        ***REMOVED***command_prefix***REMOVED***bookmark edit <id> [new name] [new timestamp]
        ///|Listing
        ***REMOVED***command_prefix***REMOVED***bookmark list [mine]
        ///|Removal
        ***REMOVED***command_prefix***REMOVED***bookmark remove <id | name>
        """
        if len(leftover_args) > 0:
            arg = leftover_args[0].lower()
            if arg in ["list", "showall"]:
                em = Embed(title="Bookmarks")
                bookmarks = bookmark.all_bookmarks

                if "mine" in leftover_args:
                    bookmarks = filter(
                        lambda x: bookmark.get_bookmark(
                            x)["author_id"] == author.id,
                        bookmarks)

                for bm in bookmarks:
                    bm_name = bm["name"]
                    bm_author = self.get_global_user(
                        bm["author_id"]).display_name
                    bm_timestamp = to_timestamp(bm["timestamp"])
                    bm_id = bm["id"]
                    t = "*****REMOVED******REMOVED*****".format(bm_name)
                    v = "`***REMOVED******REMOVED***` starting at `***REMOVED******REMOVED***` *by* *****REMOVED******REMOVED*****".format(
                        bm_id, bm_timestamp, bm_author)
                    em.add_field(name=t, value=v, inline=False)
                return Response(embed=em)
            elif arg in ["remove", "delete"]:
                if len(leftover_args) < 2:
                    return Response("Please provide an id or a name")
                bm = bookmark.get_bookmark(" ".join(leftover_args[1:]))
                if not bm:
                    return Response("Didn't find a bookmark with that query")
                if bookmark.remove_bookmark(bm["id"]):
                    return Response("Removed bookmark `***REMOVED******REMOVED***`".format(bm["name"]))
                else:
                    return Response("Something went wrong")
            elif arg in ["edit", "change"]:
                if len(leftover_args) < 2:
                    return Response("Please provide an id")

                bm_id = leftover_args[1]
                if bm_id not in bookmark:
                    return Response(
                        "No bookmark with id `***REMOVED******REMOVED***` found".format(bm_id))

                if len(leftover_args) < 3:
                    return Response(
                        "Please also specify what you want to change")

                new_timestamp = parse_timestamp(leftover_args[-1])
                if new_timestamp is not None:  # 0 evaluates to false so I need to check this oldschool-like
                    new_name = " ".join(
                        leftover_args[2:-1]) if len(leftover_args) > 3 else None
                else:
                    new_name = " ".join(leftover_args[2:])

                if bookmark.edit_bookmark(bm_id, new_name, new_timestamp):
                    return Response(
                        "Successfully edited bookmark `***REMOVED******REMOVED***`".format(bm_id))
                else:
                    return Response("Something went wrong while editing `***REMOVED******REMOVED***`".
                                    format(bm_id))
            else:
                bm = bookmark.get_bookmark(" ".join(leftover_args))
                if bm:
                    player.playlist._add_entry(
                        URLPlaylistEntry.from_dict(player.playlist, bm[
                            "entry"]))
                    return Response("Loaded bookmark `***REMOVED***0***REMOVED***` by *****REMOVED***1***REMOVED*****".
                                    format(bm["name"],
                                           self.get_global_user(
                                               bm["author_id"]).display_name))
                else:
                    bm_timestamp = player.progress
                    bm_name = None
                    if len(leftover_args) > 1:
                        timestamp = parse_timestamp(leftover_args[-1])
                        if timestamp:
                            bm_timestamp = timestamp
                        bm_name = " ".join(
                            leftover_args[:-1]) if timestamp else " ".join(
                                leftover_args)
                    else:
                        timestamp = parse_timestamp(leftover_args[-1])
                        if timestamp:
                            bm_timestamp = timestamp
                        else:
                            bm_name = " ".join(leftover_args)

                    id = bookmark.add_bookmark(
                        player.current_entry, bm_timestamp, author.id, bm_name)
                    return Response(
                        "Created a new bookmark with the id `***REMOVED***0***REMOVED***` (\"***REMOVED***2***REMOVED***\", `***REMOVED***3***REMOVED***`)\nUse `***REMOVED***1***REMOVED***bookmark ***REMOVED***0***REMOVED***` to load it ".
                        format(id, self.config.command_prefix, bm_name,
                               to_timestamp(bm_timestamp)))

        else:
            if player.current_entry:
                id = bookmark.add_bookmark(player.current_entry,
                                           player.progress, author.id)
                return Response(
                    "Created a new bookmark with the id `***REMOVED***0***REMOVED***`\nUse `***REMOVED***1***REMOVED***bookmark ***REMOVED***0***REMOVED***` to load it ".
                    format(id, self.config.command_prefix))
            else:
                return await self.cmd_bookmark(author, player, [
                    "list",
                ])

    @block_user
    @command_info("2.0.3", 1486054560, ***REMOVED***
        "3.7.2": (1498252803, "no arguments provided crash Fixed")
    ***REMOVED***)
    async def cmd_random(self, channel, author, leftover_args):
        """
        ///|Basic
        `***REMOVED***command_prefix***REMOVED***random <item1>, <item2>, [item3], [item4]`
        ///|Use an existing set
        `***REMOVED***command_prefix***REMOVED***random <setname>`
        ///|List all the existing sets
        `***REMOVED***command_prefix***REMOVED***random list`
        ///|Creation
        `***REMOVED***command_prefix***REMOVED***random create <name>, <option1>, <option2>, [option3], [option4]`
        ///|Editing
        `***REMOVED***command_prefix***REMOVED***random edit <name>, [add | remove | replace], <item> [, item2, item3]`
        ///|Removal
        `***REMOVED***command_prefix***REMOVED***random remove <name>`
        ///|Explanation
        Choose a random item out of a list or use a pre-defined list.
        """

        if not leftover_args:
            return Response("Why u gotta be stupid?")

        items = [x.strip()
                 for x in " ".join(leftover_args).split(",") if x is not ""]

        if items[0].split()[0].lower().strip() == "create":
            if len(items) < 2:
                return Response(
                    "Can't create a set with the given arguments",
                    delete_after=20)

            set_name = "_".join(items[0].split()[1:]).lower().strip()
            set_items = items[1:]
            if self.random_sets.create_set(set_name, set_items):
                return Response(
                    "Created set *****REMOVED***0***REMOVED*****\nUse `***REMOVED***1***REMOVED***random ***REMOVED***0***REMOVED***` to use it!".format(
                        set_name, self.config.command_prefix),
                    delete_after=60)
            else:
                return Response(
                    "OMG, shit went bad quickly! Everything's burning!\nDUCK there he goes again, the dragon's coming. Eat HIM not me. PLEEEEEEEEEEEEEASE!"
                )
        elif items[0].split()[0].lower().strip() == "list":
            return_string = ""
            for s in self.random_sets.get_sets():
                return_string += "*****REMOVED******REMOVED*****\n```\n***REMOVED******REMOVED***```\n\n".format(
                    s[0], ", ".join(s[1]))

            return Response(return_string)
        elif items[0].split()[0].lower().strip() == "edit":
            if len(items[0].split()) < 2:
                return Response(
                    "Please provide the name of the list you wish to edit!",
                    delete_after=20)

            set_name = "_".join(items[0].split()[1:]).lower().strip()

            existing_items = self.random_sets.get_set(set_name)
            if existing_items is None:
                return Response("This set does not exist!")

            edit_mode = items[1].strip().lower() if len(items) > 1 else None
            if edit_mode is None:
                return Response(
                    "You need to provide the way you want to edit the list",
                    delete_after=20)

            if len(items) < 3:
                return Response(
                    "You have to specify the items you want to add/remove or set as the new items"
                )

            if edit_mode == "add":
                for option in items[2:]:
                    self.random_sets.add_option(set_name, option)
            elif edit_mode == "remove":
                for option in items[2:]:
                    self.random_sets.remove_option(set_name, option)
            elif edit_mode == "replace":
                self.random_sets.replace_options(set_name, items[2:])
            else:
                return Response(
                    "This is not a valid edit mode!")

            return Response("Edited your set!")
        elif items[0].split()[0].lower().strip() == "remove":
            set_name = "_".join(items[0].split()[1:]).lower().strip()
            set_items = items[1:]
            res = self.random_sets.remove_set(set_name, set_items)
            if res:
                return Response("Removed set!")
            elif res is None:
                return Response("No such set!")
            else:
                return Response(
                    "OMG, shit went bad quickly! Everything's burning!\nDUCK there he goes again, the dragon's coming. Eat HIM not me. PLEEEEEEEEEEEEEASE!"
                )

        if len(items) <= 0 or items is None:
            return Response(
                "Is your name \"***REMOVED***0***REMOVED***\" by any chance?\n(This is not how this command works. Use `***REMOVED***1***REMOVED***help random` to find out how not to be a stupid *****REMOVED***0***REMOVED***** anymore)".
                format(author.name, self.config.command_prefix),
                delete_after=30)

        if len(items) <= 1:
            # return Response("Only you could use `***REMOVED***1***REMOVED***random` for one item...
            # Well done, ***REMOVED***0***REMOVED***!".format(author.name, self.config.command_prefix),
            # delete_after=30)

            query = "_".join(items[0].split())
            items = self.random_sets.get_set(query.lower().strip())
            if items is None:
                return Response("Something went wrong")

        await self.safe_send_message(channel,
                                     "I choose **" + choice(items) + "**")
